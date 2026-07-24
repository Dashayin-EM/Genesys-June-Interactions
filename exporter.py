"""
exporter.py
-----------
Exports the full Genesys June interactions analysis to a formatted Excel workbook.

Tabs produced:
  1. Summary            – Headline KPIs
  2. Channel Performance– Per media type breakdown (Voice, Email, WhatsApp) (NEW)
  3. Errors             – Error Code frequency table
  4. Daily Trend        – Day-by-day interactions / errors / abandons / AHT
  5. Hourly Trend       – Hour-by-hour counts
  6. Queue Performance  – Per-queue breakdown (all queues)
  7. Agent Performance  – Per-agent breakdown (all agents)
  8. Flow Performance   – Per-flow breakdown (all flows)
  9. IVR & Outcomes     – Exits, Segments, and Authentication stats
 10. Wrap-up Codes      – Wrap-up code frequencies
 11. Raw Error Records  – Every row with an error code or system-error disconnect

Usage:
    from exporter import export_to_excel
    export_to_excel(df, "reports/june_2026_genesys_report.xlsx")
"""

import os
import html
import pandas as pd
import numpy as np
import config as cfg
from utils import format_seconds

try:
    from openpyxl import load_workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


# ── Colour palette & Formatting ─────────────────────────────────────────────
HEADER_FILL   = PatternFill("solid", fgColor="1A3C5E") if HAS_OPENPYXL else None
ALT_ROW_FILL  = PatternFill("solid", fgColor="F4F8FC") if HAS_OPENPYXL else None
ERROR_FILL    = PatternFill("solid", fgColor="FFDADA") if HAS_OPENPYXL else None
WARN_FILL     = PatternFill("solid", fgColor="FFF3CD") if HAS_OPENPYXL else None
HEADER_FONT   = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10) if HAS_OPENPYXL else None
BOLD_FONT     = Font(bold=True, name="Segoe UI", size=10) if HAS_OPENPYXL else None
NORMAL_FONT   = Font(name="Segoe UI", size=10) if HAS_OPENPYXL else None
THIN_BORDER   = Border(bottom=Side(style="thin", color="C5D8ED")) if HAS_OPENPYXL else None

# Excel specific number formats
PCT_FORMAT = '0.00%'
NUM_FORMAT = '#,##0'


def _apply_header(ws, row=1):
    """Style the header row and add auto-filters."""
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 28
    # Add auto-filter to the table dimensions
    ws.auto_filter.ref = ws.dimensions


def _apply_table(ws, start_row=2, alt_col=None, error_col=None):
    """Apply alternating row shading, borders, and optional error highlighting."""
    for i, row in enumerate(ws.iter_rows(min_row=start_row)):
        fill = ALT_ROW_FILL if i % 2 == 0 else None
        for cell in row:
            if cell.font is None or not cell.font.bold:
                cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            if fill:
                cell.fill = fill
            
            # Format numbers and percentages automatically
            if isinstance(cell.value, (int, float)):
                if 0 < cell.value < 1 and 'rate' not in str(ws.cell(row=1, column=cell.column).value).lower(): 
                    # If it's a small decimal, treat it as a percentage unless it's a raw rate
                    pass 
                elif 'rate' in str(ws.cell(row=1, column=cell.column).value).lower() or '%' in str(ws.cell(row=1, column=cell.column).value):
                    cell.number_format = PCT_FORMAT
                    if cell.value > 1: # Adjust if percentages are 0-100 instead of 0-1
                        cell.value = cell.value / 100
                else:
                    cell.number_format = NUM_FORMAT

        # Red fill on error-column cells that are non-blank
        if error_col is not None:
            cell = row[error_col - 1]
            if cell.value and str(cell.value).strip() and str(cell.value) != '0':
                cell.fill = ERROR_FILL


def _autofit(ws):
    """Best-effort column width sizing."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or '')))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 50)


def _freeze(ws, row=2, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)


# ── Helper: explode multi-valued column ─────────────────────────────────────
def _explode(df, col):
    if col not in df.columns:
        return pd.DataFrame()
    temp = df.copy()
    temp['_val'] = temp[col].astype(str).str.split(';')
    exp = temp.explode('_val')
    exp['_val'] = exp['_val'].str.strip()
    return exp[exp['_val'].str.len() > 0]


# ── Sheet builders ───────────────────────────────────────────────────────────

def _build_summary(ws, df):
    ws.title = "Summary"
    total = len(df)

    if 'Parsed_Timestamp' in df.columns and df['Parsed_Timestamp'].notna().any():
        start_d = df['Parsed_Timestamp'].min().strftime('%d %b %Y')
        end_d   = df['Parsed_Timestamp'].max().strftime('%d %b %Y')
    else:
        start_d = end_d = 'N/A'

    has_err_code   = df[cfg.ERROR_CODE_COL] != '' if cfg.ERROR_CODE_COL in df.columns else pd.Series(False, index=df.index)
    has_sys_disc   = df['System Error Disconnect_Clean'] > 0
    has_err_cnt    = df['Error Count_Clean'] > 0
    has_wrap_to    = df[cfg.WRAP_UP_COL].str.contains("ININ-WRAP-UP-TIMEOUT", case=False, na=False) if cfg.WRAP_UP_COL in df.columns else pd.Series(False, index=df.index)
    has_rona_num   = df['Not Responding_Clean'] > 0
    has_rona_str   = (df['Not Responding_Clean'] == 0) & (df[cfg.USERS_NOT_RESPONDING_COL] != '')
    has_rona       = has_rona_num | has_rona_str
    has_flow_disc  = df['Flow Disconnect_Clean'] > 0
    impacted       = (has_err_code | has_sys_disc | has_err_cnt | has_wrap_to | has_rona | has_flow_disc).sum()

    queue_entered  = df[(df['Total Queue_Seconds'] > 0) | df['Abandoned_Bool']]
    qe_count       = len(queue_entered)
    total_abandoned= int(df['Abandoned_Bool'].sum())
    abnd_rate_q    = (total_abandoned / qe_count) if qe_count > 0 else 0

    aht = df[df['Total Handle_Seconds'] > 0]['Total Handle_Seconds'].mean() if 'Total Handle_Seconds' in df.columns else 0
    asa = df[df['Total Queue_Seconds']  > 0]['Total Queue_Seconds'].mean()  if 'Total Queue_Seconds'  in df.columns else 0

    wrapup_tos = int(has_wrap_to.sum())
    rona_c     = int(has_rona.sum())
    wrapup_base= df[df[cfg.WRAP_UP_COL] != ''].shape[0] if cfg.WRAP_UP_COL in df.columns else 0
    wrapup_pct = (wrapup_tos / wrapup_base) if wrapup_base > 0 else 0
    avg_time_to_abandon = df.loc[df['Abandoned_Bool'], 'Time to Abandon_Seconds'].mean() if 'Time to Abandon_Seconds' in df.columns else 0

    rows = [
        ("Metric", "Value", "Notes"),
        ("── OVERVIEW ──", "", ""),
        ("Total Interactions", total, "All records in the export"),
        ("Date Range Start", start_d, ""),
        ("Date Range End", end_d, ""),
        ("Full Export Completed", int(df['Full_Export_Completed_Bool'].sum()), ""),
        ("── ERRORS & FAILURES ──", "", ""),
        ("Unique Interactions Impacted", impacted, f"{impacted/total*100:.2f}% of total"),
        ("  - System Error Disconnects", int(df['System Error Disconnect_Clean'].sum()), "Categories overlap; see note in plan"),
        ("  - Flow Disconnects (mid-flow)", int(df['Flow Disconnect_Clean'].sum()), ""),
        ("  - All Flow Disconnects (incl. IVR)", int(df['All Flow Disconnect_Clean'].sum()), "Superset of Flow Disconnects"),
        ("  - Customer Disconnects", int(df['Customer Disconnect_Clean'].sum()), ""),
        ("  - Customer Short Disconnects", int(df['Customer Short Disconnect_Clean'].sum()), ""),
        ("  - Outcome Failures", int(df['Outcome Failure_Clean'].sum()), ""),
        ("  - Wrap-up Timeouts", wrapup_tos, f"{wrapup_pct*100:.2f}% of wrapup interactions"),
        ("  - Agent RONA (Ring-No-Answer)", rona_c, f"{rona_c/total*100:.2f}% of total"),
        ("Total Error Count Sum (from CSV)", int(df['Error Count_Clean'].sum()), "Raw sum; one interaction can have multiple"),
        ("── ABANDONMENT ──", "", ""),
        ("Total Abandoned", total_abandoned, ""),
        ("Abandon Rate (vs all interactions)", total_abandoned/total if total > 0 else 0, ""),
        ("Abandon Rate (queue-based)", abnd_rate_q, f"Out of {qe_count:,} queue-entered interactions"),
        ("Abandoned in Queue", int(df['Abandoned_in_Queue_Bool'].sum()), ""),
        ("Avg Time to Abandon", format_seconds(avg_time_to_abandon), "Abandoned interactions only"),
        ("── PERFORMANCE ──", "", ""),
        ("Avg Handle Time AHT (answered only)", format_seconds(aht), "Excludes unanswered (0s) rows"),
        ("Avg Queue Wait ASA (queued only)", format_seconds(asa), "Excludes non-queued (0s) rows"),
        ("Avg IVR Duration", format_seconds(df[df['Total IVR_Seconds'] > 0]['Total IVR_Seconds'].mean()) if 'Total IVR_Seconds' in df.columns else 'N/A', ""),
        ("Avg Hold Time", format_seconds(df[df['Total Hold_Seconds'] > 0]['Total Hold_Seconds'].mean()) if 'Total Hold_Seconds' in df.columns else 'N/A', ""),
        ("── CHANNEL ──", "", ""),
        ("Supervisor Barge-Ins", int(df['Barged_In_Bool'].sum()), ""),
        ("Total Transfers", int(df['Transfers_Clean'].sum()), ""),
        ("Interactions Transferred (%)", (df['Transfers_Clean'] > 0).sum()/total if total > 0 else 0, ""),
    ]

    for r_idx, row_data in enumerate(rows, start=1):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            # Formatting for percentage rows in summary
            if r_idx > 1 and c_idx == 2 and isinstance(val, float) and val <= 1.0:
                cell.number_format = PCT_FORMAT
            elif r_idx > 1 and c_idx == 2 and isinstance(val, (int, float)):
                cell.number_format = NUM_FORMAT

    _apply_header(ws, row=1)
    ws.auto_filter.ref = None # Remove filter from summary page for aesthetics

    for r_idx, row_data in enumerate(rows, start=1):
        if str(row_data[0]).startswith("──"):
            for c_idx in range(1, 4):
                cell = ws.cell(row=r_idx, column=c_idx)
                cell.font = Font(bold=True, name="Segoe UI", size=10, color="1A3C5E")
                cell.fill = PatternFill("solid", fgColor="E8F0F8")
        elif r_idx > 1:
            for c_idx in range(1, 4):
                cell = ws.cell(row=r_idx, column=c_idx)
                if r_idx % 2 == 0:
                    cell.fill = ALT_ROW_FILL
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER

    _autofit(ws)
    _freeze(ws)


def _build_media_type(ws, df):
    ws.title = "Channel Performance"
    ws.append(["Media Type", "Interactions", "Abandoned", "Abandon Rate %",
               "Errors", "Avg Handle Time", "Avg Queue Wait", "Avg Hold Time"])
    _apply_header(ws)

    if cfg.MEDIA_TYPE_COL in df.columns:
        m = df.groupby(cfg.MEDIA_TYPE_COL).agg(
            Interactions=('Abandoned_Bool', 'count'),
            Abandoned=('Abandoned_Bool', 'sum'),
            Errors=('Error Count_Clean', 'sum'),
            AHT=('Total Handle_Seconds', lambda x: x[x > 0].mean() if (x > 0).any() else 0),
            ASA=('Total Queue_Seconds',  lambda x: x[x > 0].mean() if (x > 0).any() else 0),
            Hold=('Total Hold_Seconds',  lambda x: x[x > 0].mean() if (x > 0).any() else 0)
        ).sort_values('Interactions', ascending=False)
        m['Abnd_Rate'] = m['Abandoned'] / m['Interactions']
        
        for name, row in m.iterrows():
            media_name = str(name).strip().title() if str(name).strip() else "Unknown"
            ws.append([media_name, int(row['Interactions']), int(row['Abandoned']),
                       row['Abnd_Rate'], int(row['Errors']),
                       format_seconds(row['AHT']), format_seconds(row['ASA']), format_seconds(row['Hold'])])

    _apply_table(ws, start_row=2)
    _autofit(ws)
    _freeze(ws)


def _build_errors(ws, df):
    ws.title = "Errors"
    total = len(df)
    err_col = cfg.ERROR_CODE_COL

    ws.append(["Error Code", "Occurrences", "% of Total Interactions"])
    _apply_header(ws)

    if err_col in df.columns:
        err_df = df[df[err_col] != '']
        counts = err_df[err_col].value_counts().reset_index()
        counts.columns = ['Error Code', 'Count']
        counts['Pct'] = counts['Count'] / total
        for _, row in counts.iterrows():
            ws.append([row['Error Code'], int(row['Count']), row['Pct']])

    _apply_table(ws, start_row=2, error_col=1)
    _autofit(ws)
    _freeze(ws)


def _build_daily(ws, df):
    ws.title = "Daily Trend"
    ws.append(["Date", "Interactions", "Errors", "Abandons", "Abandon Rate %", "Avg Handle Time"])
    _apply_header(ws)

    if 'DateOnly' in df.columns:
        daily = df.groupby('DateOnly').agg(
            Interactions=('Abandoned_Bool', 'count'),
            Errors=('Error Count_Clean', 'sum'),
            Abandoned=('Abandoned_Bool', 'sum'),
            AHT=('Total Handle_Seconds', lambda x: x[x > 0].mean() if (x > 0).any() else 0)
        )
        daily['Abandon_Rate'] = daily['Abandoned'] / daily['Interactions']
        for date_val, row in daily.iterrows():
            ws.append([
                str(date_val),
                int(row['Interactions']),
                int(row['Errors']),
                int(row['Abandoned']),
                row['Abandon_Rate'],
                format_seconds(row['AHT'])
            ])

    _apply_table(ws, start_row=2)
    _autofit(ws)
    _freeze(ws)


def _build_hourly(ws, df):
    ws.title = "Hourly Trend"
    ws.append(["Hour", "Interactions", "Errors", "Abandons"])
    _apply_header(ws)

    if 'Hour' in df.columns:
        hourly = df.groupby('Hour').agg(
            Interactions=('Abandoned_Bool', 'count'),
            Errors=('Error Count_Clean', 'sum'),
            Abandoned=('Abandoned_Bool', 'sum')
        )
        for hour, row in hourly.iterrows():
            ws.append([f"{int(hour):02d}:00", int(row['Interactions']), int(row['Errors']), int(row['Abandoned'])])

    _apply_table(ws, start_row=2)
    _autofit(ws)
    _freeze(ws)


def _build_queue(ws, df):
    ws.title = "Queue Performance"
    ws.append(["Queue", "Interactions", "Abandoned", "Abandon Rate %",
               "Errors", "Avg Handle Time", "Avg Queue Wait"])
    _apply_header(ws)

    exp = _explode(df, cfg.QUEUE_COL)
    if not exp.empty:
        q = exp.groupby('_val').agg(
            Interactions=('Abandoned_Bool', 'count'),
            Abandoned=('Abandoned_Bool', 'sum'),
            Errors=('Error Count_Clean', 'sum'),
            AHT=('Total Handle_Seconds', lambda x: x[x > 0].mean() if (x > 0).any() else 0),
            ASA=('Total Queue_Seconds',  lambda x: x[x > 0].mean() if (x > 0).any() else 0)
        ).sort_values('Interactions', ascending=False)
        q['Abnd_Rate'] = q['Abandoned'] / q['Interactions']
        for name, row in q.iterrows():
            ws.append([name, int(row['Interactions']), int(row['Abandoned']),
                       row['Abnd_Rate'], int(row['Errors']),
                       format_seconds(row['AHT']), format_seconds(row['ASA'])])

    _apply_table(ws, start_row=2)
    _autofit(ws)
    _freeze(ws)


def _build_agent(ws, df):
    ws.title = "Agent Performance"
    ws.append(["Agent", "Interactions", "Errors", "Wrap-up Timeouts",
               "RONA Events", "Transfers", "Avg Handle Time", "Avg Hold Time"])
    _apply_header(ws)

    exp = _explode(df, cfg.AGENT_COL)
    if not exp.empty:
        exp['Is_Timeout'] = exp[cfg.WRAP_UP_COL].str.contains("ININ-WRAP-UP-TIMEOUT", case=False, na=False)
        exp['Is_RONA']    = (exp['Not Responding_Clean'] > 0) |                             ((exp['Not Responding_Clean'] == 0) & (exp[cfg.USERS_NOT_RESPONDING_COL] != ''))
        a = exp.groupby('_val').agg(
            Interactions=('Abandoned_Bool', 'count'),
            Errors=('Error Count_Clean', 'sum'),
            Timeouts=('Is_Timeout', 'sum'),
            RONA=('Is_RONA', 'sum'),
            Transfers=('Transfers_Clean', 'sum'),
            AHT=('Total Handle_Seconds', lambda x: x[x > 0].mean() if (x > 0).any() else 0),
            Hold=('Total Hold_Seconds',  lambda x: x[x > 0].mean() if (x > 0).any() else 0)
        ).sort_values('Interactions', ascending=False)
        for name, row in a.iterrows():
            ws.append([name, int(row['Interactions']), int(row['Errors']),
                       int(row['Timeouts']), int(row['RONA']), int(row['Transfers']),
                       format_seconds(row['AHT']), format_seconds(row['Hold'])])

    _apply_table(ws, start_row=2)
    _autofit(ws)
    _freeze(ws)


def _build_flow(ws, df):
    ws.title = "Flow Performance"
    ws.append(["Flow Name", "Interactions", "Flow Disconnects (mid-flow)", "Errors"])
    _apply_header(ws)

    exp = _explode(df, cfg.FLOW_COL)
    if not exp.empty:
        f = exp.groupby('_val').agg(
            Interactions=('Abandoned_Bool', 'count'),
            Flow_Disconnects=('Flow Disconnect_Clean', 'sum'),
            Errors=('Error Count_Clean', 'sum')
        ).sort_values('Interactions', ascending=False)
        for name, row in f.iterrows():
            ws.append([name, int(row['Interactions']),
                       int(row['Flow_Disconnects']), int(row['Errors'])])

    _apply_table(ws, start_row=2)
    _autofit(ws)
    _freeze(ws)


def _build_ivr_outcomes(ws, df):
    ws.title = "IVR & Outcomes"
    total = len(df)
    ws.append(["Category", "Metric / Path", "Occurrences", "% of Total"])
    _apply_header(ws)
    
    def append_counts(category_name, col_name, top_n=15):
        if col_name in df.columns:
            counts = df[df[col_name] != ''][col_name].value_counts().head(top_n)
            for val, count in counts.items():
                if str(val).strip():
                    ws.append([category_name, str(val), int(count), count/total])
    
    append_counts("Flow Exit Reason", cfg.FLOW_EXIT_COL)
    append_counts("Flow-Out Type", cfg.FLOW_OUT_TYPE_COL)
    append_counts("IVR Segment Path", cfg.IVR_SEGMENTS_COL)
    append_counts("Failed Outcomes", cfg.FAILED_OUTCOMES_COL)
    append_counts("Authenticated Status", cfg.AUTHENTICATED_COL)
    append_counts("Customer Journey Data", cfg.HAS_CUSTOMER_JOURNEY_DATA_COL)
    
    _apply_table(ws, start_row=2)
    _autofit(ws)
    _freeze(ws)


def _build_wrapup(ws, df):
    ws.title = "Wrap-up Codes"
    total = len(df)
    ws.append(["Wrap-up Code", "Occurrences", "% of Total Interactions"])
    _apply_header(ws)

    exp = _explode(df, cfg.WRAP_UP_COL)
    if not exp.empty:
        counts = exp['_val'].value_counts()
        for code, count in counts.items():
            row_fill = ERROR_FILL if 'ININ-WRAP-UP-TIMEOUT' in str(code).upper() else None
            r = ws.max_row + 1
            ws.append([code, int(count), count / total])
            if row_fill:
                for c in range(1, 4):
                    ws.cell(row=r, column=c).fill = row_fill

    _apply_table(ws, start_row=2)
    _autofit(ws)
    _freeze(ws)


def _build_raw_errors(ws, df):
    ws.title = "Raw Error Records"

    has_err_code  = (df[cfg.ERROR_CODE_COL] != '') if cfg.ERROR_CODE_COL in df.columns else pd.Series(False, index=df.index)
    has_sys_disc  = df['System Error Disconnect_Clean'] > 0
    has_err_cnt   = df['Error Count_Clean'] > 0
    has_wrap_to   = df[cfg.WRAP_UP_COL].str.contains("ININ-WRAP-UP-TIMEOUT", case=False, na=False) if cfg.WRAP_UP_COL in df.columns else pd.Series(False, index=df.index)
    has_flow_disc = df['Flow Disconnect_Clean'] > 0
    
    has_rona_num  = df['Not Responding_Clean'] > 0
    has_rona_str  = (df['Not Responding_Clean'] == 0) & (df[cfg.USERS_NOT_RESPONDING_COL] != '') if cfg.USERS_NOT_RESPONDING_COL in df.columns else pd.Series(False, index=df.index)
    has_rona      = has_rona_num | has_rona_str

    mask = has_err_code | has_sys_disc | has_err_cnt | has_wrap_to | has_rona | has_flow_disc

    error_df = df[mask].copy()

    cols = [c for c in [
        cfg.TIMESTAMP_COL, 'Interaction ID', cfg.ANI_COL, cfg.DNIS_COL, cfg.REMOTE_COL, 
        cfg.MEDIA_TYPE_COL, cfg.DIRECTION_COL,
        cfg.QUEUE_COL, cfg.AGENT_COL, cfg.USERS_NOT_RESPONDING_COL, 
        cfg.FLOW_COL, cfg.FLOW_EXIT_COL, cfg.IVR_SEGMENTS_COL, cfg.FAILED_OUTCOMES_COL,
        cfg.ERROR_CODE_COL, 'Error Count_Clean',
        'System Error Disconnect_Clean', 'Flow Disconnect_Clean',
        cfg.DISCONNECT_TYPE_COL, cfg.WRAP_UP_COL,
        'Abandoned_Bool', 'Duration_Seconds', 'Total Handle_Seconds',
    ] if c in error_df.columns]

    ws.append(cols)
    _apply_header(ws)

    for _, row in error_df[cols].iterrows():
        ws.append([str(v) if isinstance(v, bool) else v for v in row.tolist()])

    err_col_idx = cols.index(cfg.ERROR_CODE_COL) + 1 if cfg.ERROR_CODE_COL in cols else None
    _apply_table(ws, start_row=2, error_col=err_col_idx)
    _autofit(ws)
    _freeze(ws)


# ── Main export function ─────────────────────────────────────────────────────

def export_to_excel(df: pd.DataFrame, output_path: str):
    """
    Write the full analysis to a formatted Excel workbook.
    """
    if not HAS_OPENPYXL:
        print("⚠️  openpyxl is not installed. Run: pip install openpyxl")
        return

    os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)

    from openpyxl import Workbook
    wb = Workbook()

    ws_summary = wb.active
    _build_summary(ws_summary, df)

    # Added the Media Type breakdown here
    _build_media_type(wb.create_sheet(), df)
    
    _build_errors(wb.create_sheet(), df)
    _build_daily(wb.create_sheet(), df)
    _build_hourly(wb.create_sheet(), df)
    _build_queue(wb.create_sheet(), df)
    _build_agent(wb.create_sheet(), df)
    _build_flow(wb.create_sheet(), df)
    _build_ivr_outcomes(wb.create_sheet(), df) 
    _build_wrapup(wb.create_sheet(), df)
    _build_raw_errors(wb.create_sheet(), df)

    wb.save(output_path)
    print(f"📊 Excel report saved → {output_path}")
    print(f"   Tabs: {', '.join(ws.title for ws in wb.worksheets)}")


def append_terminal_details(xlsx_path: str, html_path: str, terminal_output: str):
    """Parse raw terminal text and append it as sleek, collapsible HTML accordions."""
    import os
    import html
    import re
    
    # 1. Write the terminal log to the Excel file
    if HAS_OPENPYXL:
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment
        try:
            wb = load_workbook(xlsx_path)
            if "Terminal Detail" in wb.sheetnames:
                del wb["Terminal Detail"]
            ws = wb.create_sheet("Terminal Detail")
            ws["A1"] = "Genesys Runtime Analysis"
            ws["A2"] = "Exact console output from this report run"
            ws.column_dimensions["A"].width = 120
            for row_number, line in enumerate(terminal_output.splitlines(), start=4):
                ws.cell(row=row_number, column=1, value=line).alignment = Alignment(wrap_text=True, vertical="top")
            wb.save(xlsx_path)
        except Exception as e:
            print(f"Warning: Could not save terminal log to Excel: {e}")

    # 2. Parse the terminal output into Collapsible HTML Accordions
    if not os.path.exists(html_path):
        return
        
    with open(html_path, "r", encoding="utf-8") as f:
        html_report = f.read()

    # Start the section container
    html_blocks = [
        "<div class='section-title' style='margin-top:4rem;'>Comprehensive System Report (Click to Expand)</div>",
        "<div class='accordion-container' style='display: flex; flex-direction: column; gap: 1rem; margin-bottom: 3rem;'>"
    ]
    
    lines = terminal_output.split('\n')
    in_table = False
    in_section = False
    table_html = ""
    
    for line in lines:
        line_str = line.strip()
        
        # Skip raw ASCII borders
        if not line_str or set(line_str) == {'-'} or set(line_str) == {'='}:
            continue
            
        # Detect numbered section headers (e.g., "1. DATA INTEGRITY")
        if re.match(r'^\d+\.\s+', line_str):
            if in_table:
                html_blocks.append(table_html + "</tbody></table></div>")
                in_table = False
                
            # Close the previous accordion if one is open
            if in_section:
                html_blocks.append("</div></details>")
                
            # Start a new collapsible accordion
            html_blocks.append(f"""
            <details style='background: var(--surface); border: 1px solid var(--accent); border-radius: var(--radius); overflow: hidden; box-shadow: 0 2px 4px rgba(0,0,0,0.02);'>
                <summary style='padding: 1.2rem 1.5rem; font-weight: 600; color: var(--primary); font-size: 1.05rem; cursor: pointer; background: #f8fafc; list-style: none; display: flex; justify-content: space-between; align-items: center; transition: background 0.2s;'>
                    <span>{html.escape(line_str)}</span>
                    <span style='color: var(--secondary); font-size: 0.8em;'>▼ Expand</span>
                </summary>
                <div style='padding: 1.5rem; border-top: 1px solid var(--accent);'>
            """)
            in_section = True
            continue
            
        # Detect ASCII tables and convert to proper HTML tables
        if '|' in line:
            if not in_table:
                in_table = True
                headers = [h.strip() for h in line.split('|') if h.strip()]
                th_html = "".join(f"<th>{html.escape(h)}</th>" for h in headers)
                table_html = f"<div class='table-wrapper' style='margin-bottom: 1.5rem; overflow-x: auto;'><table class='dataTable display' style='width: 100%; text-align: left;'><thead><tr style='background: var(--primary); color: white;'>{th_html}</tr></thead><tbody>"
            else:
                cells = [c.strip() for c in line.split('|') if c.strip()]
                if cells:
                    td_html = "".join(f"<td style='padding: 0.8rem; border-bottom: 1px solid var(--accent);'>{html.escape(c)}</td>" for c in cells)
                    table_html += f"<tr style='background: var(--surface);'>{td_html}</tr>"
            continue
            
        # Close HTML table if the ASCII table ends
        if in_table:
            html_blocks.append(table_html + "</tbody></table></div>")
            in_table = False
            
        # Format bullet points and metric lists
        if line.startswith('   -') or line.startswith('   └') or line.startswith(' 🚨') or line.startswith(' 🚪') or line.startswith(' 📴'):
            html_blocks.append(f"<div class='stat-row' style='padding: 0.5rem 1rem; border-top: none; border-bottom: 1px solid var(--accent); justify-content: flex-start; color: var(--text-main); font-weight: 500;'>{html.escape(line_str)}</div>")
        else:
            # Regular text paragraphs
            html_blocks.append(f"<p style='font-size: 0.95rem; margin-bottom: 0.5rem; color: var(--text-muted);'>{html.escape(line_str)}</p>")
            
    # Cleanup any open blocks at the end of the loop
    if in_table:
        html_blocks.append(table_html + "</tbody></table></div>")
    if in_section:
        html_blocks.append("</div></details>")
        
    html_blocks.append("</div>") # Close accordion container
    
    parsed_html = "".join(html_blocks)
    
    # Inject the HTML right before the end of the document
    if "</body>" in html_report:
        html_report = html_report.replace("</body>", parsed_html + "\n</body>", 1)
        
    with open(html_path, "w", encoding="utf-8") as f:
        f.write(html_report)